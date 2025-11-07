from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import CustomAuthenticationForm
from .models import Proveedor, Producto, Desempeno
from .forms import ProveedorForm, EditarPerfilForm, AgregarProductoForm
from django.contrib.auth.models import User
from django.db.models import Avg
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.decorators import permission_required
from django.contrib.auth import update_session_auth_hash
import random
from django.core.mail import send_mail
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import logging

logger = logging.getLogger(__name__)

@login_required
def exportar_excel(request):
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen de Datos"

    # 🎨 Estilos
    bold_center = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # --- 🟩 SECCIÓN PROVEEDORES ---
    ws.merge_cells("A1:D1")
    ws["A1"] = "LISTA DE PROVEEDORES"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].alignment = center_align
    ws["A1"].fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

    headers_proveedores = ["Nombre", "RUT", "Contacto", "Email"]
    ws.append(headers_proveedores)

    for col_num, header in enumerate(headers_proveedores, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = bold_center
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    proveedores = Proveedor.objects.all()
    for p in proveedores:
        ws.append([p.nombre, p.rut, p.contacto, getattr(p, 'email', '-')])

    # Ajustar anchos de columna (sin error de merged cells)
    for i, column_cells in enumerate(ws.columns, 1):
        length = 0
        for cell in column_cells:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = length + 3

    # --- 🟨 ESPACIO ---
    fila_actual = ws.max_row + 2

    # --- 🟦 SECCIÓN PRODUCTOS ---
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=3)
    ws.cell(row=fila_actual, column=1).value = "LISTA DE PRODUCTOS"
    ws.cell(row=fila_actual, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=fila_actual, column=1).alignment = center_align
    ws.cell(row=fila_actual, column=1).fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")

    fila_actual += 1
    headers_productos = ["Tipo de producto", "Producto", "Proveedor"]
    for col_num, header in enumerate(headers_productos, 1):
        cell = ws.cell(row=fila_actual, column=col_num)
        cell.value = header
        cell.font = bold_center
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    productos = Producto.objects.all()
    for prod in productos:
        fila_actual += 1
        ws.append([
            prod.tipo_producto if hasattr(prod, 'tipo_producto') else "-",
            prod.nombre,
            prod.proveedor.nombre if prod.proveedor else "-"
        ])

    # --- 🟧 ESPACIO ---
    fila_actual = ws.max_row + 2

    # --- 🟪 SECCIÓN USUARIOS ---
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=4)
    ws.cell(row=fila_actual, column=1).value = "LISTA DE USUARIOS"
    ws.cell(row=fila_actual, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=fila_actual, column=1).alignment = center_align
    ws.cell(row=fila_actual, column=1).fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

    fila_actual += 1
    headers_usuarios = ["Usuario", "Nombre completo", "Email", "Es staff"]
    for col_num, header in enumerate(headers_usuarios, 1):
        cell = ws.cell(row=fila_actual, column=col_num)
        cell.value = header
        cell.font = bold_center
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    usuarios = User.objects.all()
    for user in usuarios:
        fila_actual += 1
        ws.append([
            user.username,
            f"{user.first_name} {user.last_name}",
            user.email,
            "Sí" if user.is_staff else "No"
        ])

    # Ajustar ancho de columnas finales
    for i, column_cells in enumerate(ws.columns, 1):
        length = 0
        for cell in column_cells:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = length + 3

    # --- 📤 EXPORTAR ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Datos_Secreto_Gelateria.xlsx"'
    wb.save(response)
    return response

def cambiar_contrasena(request):
    user_id = request.session.get('usuario_recuperacion')
    if not user_id:
        messages.error(request, "No hay usuario para cambiar la contraseña.")
        return redirect('recordar_contrasena')

    user = User.objects.get(id=user_id)

    if request.method == 'POST':
        nueva_pass = request.POST.get('password1')
        confirmar_pass = request.POST.get('password2')
        if nueva_pass != confirmar_pass:
            messages.error(request, "Las contraseñas no coinciden.")
        else:
            user.set_password(nueva_pass)
            user.save()
            # Limpiar la sesión
            request.session.pop('codigo_recuperacion', None)
            request.session.pop('usuario_recuperacion', None)
            messages.success(request, "Contraseña cambiada correctamente.")
            return redirect('login')

    return render(request, 'proveedores/cambiar_contrasena.html')

def verificar_codigo(request):
    if request.method == 'POST':
        codigo_ingresado = request.POST.get('codigo')
        codigo_guardado = request.session.get('codigo_recuperacion')
        user_id = request.session.get('usuario_recuperacion')

        if not codigo_guardado or not user_id:
            messages.error(request, "No hay un código para verificar. Intenta de nuevo.")
            return redirect('recordar_contrasena')

        if str(codigo_ingresado) == str(codigo_guardado):
            # Código correcto, puedes redirigir a página de cambio de contraseña
            messages.success(request, "Código verificado. Ahora puedes cambiar tu contraseña.")
            return redirect('cambiar_contrasena')  # Crear esta vista después
        else:
            messages.error(request, "Código incorrecto, intenta nuevamente.")

    return render(request, 'proveedores/verificar_codigo.html')

def recordar_contrasena(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            codigo = random.randint(100000, 999999)
            # Guardar el código en sesión por ahora (para ejemplo simple)
            request.session['codigo_recuperacion'] = codigo
            request.session['usuario_recuperacion'] = user.id

            send_mail(
                'Código de recuperación de contraseña',
                f'Hola {user.first_name or user.username}, tu código para recuperar la contraseña es: {codigo}',
                None,
                [email],
                fail_silently=False,
            )

            messages.success(request, 'Se ha enviado un código a tu correo.')
            return redirect('verificar_codigo')  # Crear otra vista para verificar el código
        except User.DoesNotExist:
            messages.error(request, 'No se encontró ningún usuario con ese correo.')

    return render(request, 'proveedores/recordar_contrasena.html')

@login_required
def agregar_producto(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, pk=proveedor_id)

    if request.method == 'POST':
        form = AgregarProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.proveedor = proveedor
            producto.save()

            # 🔥 Si la petición viene por AJAX, respondemos JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'producto': {
                        'nombre': producto.nombre,
                        'tipo_producto': producto.tipo_producto,
                    }
                })

            messages.success(request, "Producto agregado correctamente.")
            return redirect('detalle_proveedor', proveedor_id=proveedor.id)
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False})
            messages.error(request, "Error al agregar el producto. Revisa los campos.")
            return redirect('detalle_proveedor', proveedor_id=proveedor.id)

    return redirect('detalle_proveedor', proveedor_id=proveedor.id)


@login_required
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    proveedor_id = producto.proveedor.id  # Para redirigir después
    producto.delete()
    messages.success(request, f'Producto "{producto.nombre}" eliminado correctamente.')
    return redirect('detalle_proveedor', proveedor_id=proveedor_id)

@login_required
def editar_perfil(request):
    user = request.user

    if request.method == "POST":
        form = EditarPerfilForm(request.POST, instance=user)
        if form.is_valid():
            # Actualizamos los campos normales
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data['email']
            user.username = form.cleaned_data['email']  # ⚡ Sincronizar username con email

            # Actualizar contraseña si se ingresó
            nueva_pass = form.cleaned_data.get('password')
            if nueva_pass:
                user.set_password(nueva_pass)
                update_session_auth_hash(request, user)  # Mantener sesión activa

            user.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return redirect('editar_perfil')
        else:
            messages.error(request, "Corrige los errores en el formulario.")
    else:
        form = EditarPerfilForm(instance=user)

    return render(request, 'proveedores/editar_perfil.html', {'form': form})

@login_required
@permission_required('proveedores.change_proveedor', raise_exception=True)  # Solo editores
def editor_view(request):
    # Filtros
    search_proveedor = request.GET.get('search_proveedor', '').strip()
    proveedor_id = request.GET.get('proveedor', '').strip()
    tipo_producto = request.GET.get('tipo_producto', '').strip()
    nombre_producto = request.GET.get('nombre', '').strip()
    orden = request.GET.get('orden', '')

    # Listados iniciales
    proveedores_list = Proveedor.objects.all().order_by('nombre')
    productos_list = Producto.objects.select_related('proveedor').all().order_by('nombre')

    # Filtrar proveedores
    if search_proveedor:
        proveedores_list = proveedores_list.filter(
            Q(nombre__icontains=search_proveedor) |
            Q(rut__icontains=search_proveedor) |
            Q(contacto__icontains=search_proveedor)
        )

    # Filtrar productos
    if proveedor_id:
        try:
            pid = int(proveedor_id)
            productos_list = productos_list.filter(proveedor__id=pid)
        except ValueError:
            pass

    if tipo_producto:
        productos_list = productos_list.filter(tipo_producto=tipo_producto)

    if nombre_producto:
        productos_list = productos_list.filter(nombre__icontains=nombre_producto)

    # Ordenar productos
    if orden == 'asc':
        productos_list = productos_list.order_by('nombre')
    elif orden == 'desc':
        productos_list = productos_list.order_by('-nombre')

    # Paginación (10 por página)
    proveedores_paginator = Paginator(proveedores_list, 10)
    productos_paginator = Paginator(productos_list, 10)

    proveedores_page_number = request.GET.get('proveedores_page')
    productos_page_number = request.GET.get('productos_page')

    proveedores = proveedores_paginator.get_page(proveedores_page_number)
    productos = productos_paginator.get_page(productos_page_number)

    # Tipos de producto distintos (ahora vienen de Producto)
    tipos_producto_distinct = Producto.objects.values_list('tipo_producto', flat=True).distinct()

    context = {
        'proveedores': proveedores,
        'productos': productos,
        'tipos_producto': tipos_producto_distinct,
        'search_proveedor': search_proveedor
    }

    return render(request, 'proveedores/dashboard_editor.html', context)

@permission_required('proveedores.view_proveedor', raise_exception=True)
def solo_vista(request):
    # Obtener datos base
    proveedores = Proveedor.objects.all()
    productos = Producto.objects.select_related('proveedor').all()
    tipos_producto = Producto.objects.values_list('tipo_producto', flat=True).distinct()

    # --- Filtros de búsqueda ---
    search_proveedor = request.GET.get('search_proveedor', '').strip()
    tipo_producto = request.GET.get('tipo_producto', '').strip()
    proveedor_id = request.GET.get('proveedor', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    orden = request.GET.get('orden', '').strip()

    # Filtro: búsqueda por nombre o RUT o contacto del proveedor
    if search_proveedor:
        proveedores = proveedores.filter(
            nombre__icontains=search_proveedor
        ) | proveedores.filter(
            rut__icontains=search_proveedor
        ) | proveedores.filter(
            contacto__icontains=search_proveedor
        )

    # Filtro: productos por tipo (ahora en Producto)
    if tipo_producto:
        productos = productos.filter(tipo_producto=tipo_producto)

    # Filtro: productos por proveedor
    if proveedor_id:
        try:
            productos = productos.filter(proveedor__id=int(proveedor_id))
        except ValueError:
            pass

    # Filtro: productos por nombre
    if nombre:
        productos = productos.filter(nombre__icontains=nombre)

    # Ordenar A-Z o Z-A
    if orden == 'asc':
        productos = productos.order_by('nombre')
        proveedores = proveedores.order_by('nombre')
    elif orden == 'desc':
        productos = productos.order_by('-nombre')
        proveedores = proveedores.order_by('-nombre')

    # --- Paginadores ---
    proveedor_page = request.GET.get('proveedores_page', 1)
    producto_page = request.GET.get('productos_page', 1)

    proveedor_paginator = Paginator(proveedores, 10)  # 🔥 Ahora muestra 10
    producto_paginator = Paginator(productos, 10)    # 🔥 Ahora muestra 10

    proveedores = proveedor_paginator.get_page(proveedor_page)
    productos = producto_paginator.get_page(producto_page)

    # --- Contexto ---
    context = {
        'proveedores': proveedores,
        'productos': productos,
        'tipos_producto': tipos_producto,
        'search_proveedor': search_proveedor,
    }
    return render(request, 'proveedores/solo_vista.html', context)

def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    proveedor.delete()
    messages.success(request, f'Proveedor "{proveedor.nombre}" eliminado correctamente.')
    return redirect('dashboard')

@login_required
def detalle_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
    form = ProveedorForm(instance=proveedor)
    producto_form = AgregarProductoForm()  # para el modal

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor actualizado correctamente.")
            return redirect('detalle_proveedor', proveedor_id=proveedor.id)
        else:
            messages.error(request, "Error al actualizar el proveedor. Revisa los campos.")

    context = {
        'proveedor': proveedor,
        'form': form,
        'producto_form': producto_form,   # importante para que el modal tenga acceso
        'productos': proveedor.productos.all(),  # para mostrar en la tabla
    }
    return render(request, 'proveedores/detalle_proveedor.html', context)

@login_required
def gestionar_desempeno(request):
    productos = Producto.objects.select_related('proveedor').all()

    # Filtros
    tipo_producto = request.GET.get('tipo_producto', '').strip()
    proveedor_id = request.GET.get('proveedor', '').strip()
    nombre_producto = request.GET.get('nombre', '').strip()
    orden = request.GET.get('orden', '').strip()

    if tipo_producto:
        productos = productos.filter(tipo_producto=tipo_producto)
    if proveedor_id:
        try:
            pid = int(proveedor_id)
            productos = productos.filter(proveedor__id=pid)
        except ValueError:
            pass
    if nombre_producto:
        productos = productos.filter(nombre__icontains=nombre_producto)
    if orden == 'asc':
        productos = productos.order_by('nombre')
    elif orden == 'desc':
        productos = productos.order_by('-nombre')

    # Paginación: 10 por página
    paginator = Paginator(productos, 10)
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)

    proveedores = Proveedor.objects.all()
    tipos_producto = Producto.objects.values_list('tipo_producto', flat=True).distinct()

    return render(request, 'proveedores/gestionar_desempeno.html', {
        'productos': productos_page,
        'proveedores': proveedores,
        'tipos_producto': tipos_producto,
    })

def gestionar_desempeno_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    productos = proveedor.productos.all()
    desempenos_list = proveedor.desempenos.all().order_by('-fecha')

    # Paginación del historial de desempeño
    paginator = Paginator(desempenos_list, 10)  # 10 por página
    page_number = request.GET.get('desempenos_page')
    desempenos = paginator.get_page(page_number)

    # Calcular desempeño global
    if desempenos_list.exists():
        total_desempenos = desempenos_list.count()
        promedio_calidad = sum(d.calidad for d in desempenos_list) / total_desempenos
        promedio_entrega = sum(d.entrega for d in desempenos_list) / total_desempenos
        promedio_precio = sum(d.precio for d in desempenos_list) / total_desempenos
        ranking_global = (promedio_calidad + promedio_entrega + promedio_precio) / 3
    else:
        promedio_calidad = promedio_entrega = promedio_precio = ranking_global = 0

    swal_success = None
    swal_error = None

    if request.method == "POST":
        error = False
        for producto in productos:
            calidad_raw = request.POST.get(f'calidad_{producto.id}', '').strip()
            entrega_raw = request.POST.get(f'entrega_{producto.id}', '').strip()
            precio_raw = request.POST.get(f'precio_{producto.id}', '').strip()

            if not calidad_raw or not entrega_raw or not precio_raw:
                error = True
                break

            try:
                calidad = int(calidad_raw)
                entrega = int(entrega_raw)
                precio = int(precio_raw)
            except ValueError:
                error = True
                break

            if not (1 <= calidad <= 10) or not (1 <= entrega <= 10) or not (1 <= precio <= 10):
                error = True
                break

            Desempeno.objects.create(
                proveedor=proveedor,
                calidad=calidad,
                entrega=entrega,
                precio=precio
            )

        if error:
            swal_error = "Faltan campos por rellenar o los valores son inválidos (1-10)."
        else:
            swal_success = "Desempeño registrado correctamente."

    return render(request, 'proveedores/gestionar_desempeno_proveedor.html', {
        'proveedor': proveedor,
        'productos': productos,
        'desempenos': desempenos,
        'promedio_calidad': promedio_calidad,
        'promedio_entrega': promedio_entrega,
        'promedio_precio': promedio_precio,
        'ranking_global': ranking_global,
        'swal_success': swal_success,
        'swal_error': swal_error
    })

@login_required
def nuevo_proveedor(request):
    if request.method == "POST":
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()  # Guardamos el proveedor
            producto_nombre = form.cleaned_data.get('producto_nombre')
            if producto_nombre:
                Producto.objects.create(
                    nombre=producto_nombre,
                    tipo_producto='Liquido',  # Por defecto o puedes agregar select
                    proveedor=proveedor
                )

            messages.success(request, f"Proveedor '{proveedor.nombre}' agregado correctamente.")
            return redirect('nuevo_proveedor')
        else:
            messages.error(request, "Error al agregar el proveedor. Revisa los campos.")
    else:
        form = ProveedorForm()

    # Traemos todos los proveedores y paginamos
    proveedores_list = Proveedor.objects.all().order_by("nombre")
    proveedores_paginator = Paginator(proveedores_list, 10)
    proveedores_page_number = request.GET.get('page')
    proveedores = proveedores_paginator.get_page(proveedores_page_number)

    return render(request, 'proveedores/nuevo_proveedor.html', {
        'form': form,
        'proveedores': proveedores
    })


def login_view(request):
    if request.method == "POST":
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # 🔒 Redirección según permisos
            if user.is_staff:
                redirect_url = '/dashboard/'  # Admin
            elif user.has_perm('proveedores.change_proveedor'):
                redirect_url = '/dashboard_editor/'  # Editor
            elif user.has_perm('proveedores.view_proveedor'):
                redirect_url = '/solo_vista/'  # Usuario solo lectura
            else:
                redirect_url = '/'  # Default o página de inicio

            # 🔥 Mostramos mensaje SweetAlert con redirección
            return render(request, 'proveedores/login.html', {
                'form': form,
                'swal_success': f"Bienvenido {user.first_name or user.username}",
                'redirect_url': redirect_url,
            })

        else:
            # ❌ Usuario o contraseña incorrectos
            return render(request, 'proveedores/login.html', {
                'form': form,
                'swal_error': "Usuario o contraseña incorrectos"
            })
    else:
        form = CustomAuthenticationForm()

    return render(request, 'proveedores/login.html', {'form': form})

# Registro funcional
def register_view(request):
    if request.method == "POST":
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            messages.error(request, "Las contraseñas no coinciden")
        elif User.objects.filter(username=email).exists():
            messages.error(request, "El usuario ya existe")
        else:
            User.objects.create_user(
                username=email,
                first_name=nombre,
                last_name=apellido,
                email=email,
                password=password1
            )
            messages.success(request, "Usuario registrado correctamente")
            return redirect('login')

    return render(request, 'proveedores/registrarse.html')

# Dashboard protegido
@login_required
def dashboard(request):
    # --- Filtros existentes ---
    search_proveedor = request.GET.get('search_proveedor', '').strip()
    proveedor_id = request.GET.get('proveedor', '').strip()
    tipo_producto = request.GET.get('tipo_producto', '').strip()
    nombre_producto = request.GET.get('nombre', '').strip()
    orden = request.GET.get('orden', '').strip()  # 🆕 nuevo filtro de orden

    # --- Listados iniciales ---
    proveedores_list = Proveedor.objects.all()
    productos_list = Producto.objects.select_related('proveedor').all()

    # --- Filtrar proveedores por búsqueda ---
    if search_proveedor:
        proveedores_list = proveedores_list.filter(
            Q(nombre__icontains=search_proveedor) |
            Q(rut__icontains=search_proveedor) |
            Q(contacto__icontains=search_proveedor)
        )

    # --- Filtrar productos ---
    if proveedor_id:
        try:
            pid = int(proveedor_id)
            productos_list = productos_list.filter(proveedor__id=pid)
        except ValueError:
            pass

    if tipo_producto:
        productos_list = productos_list.filter(tipo_producto=tipo_producto)

    if nombre_producto:
        productos_list = productos_list.filter(nombre__icontains=nombre_producto)

    # --- Ordenar ---
    if orden == "asc":
        proveedores_list = proveedores_list.order_by("nombre")
        productos_list = productos_list.order_by("nombre")
    elif orden == "desc":
        proveedores_list = proveedores_list.order_by("-nombre")
        productos_list = productos_list.order_by("-nombre")
    else:
        proveedores_list = proveedores_list.order_by("nombre")
        productos_list = productos_list.order_by("nombre")

    # --- Paginación de proveedores y productos ---
    proveedores_paginator = Paginator(proveedores_list, 10)
    productos_paginator = Paginator(productos_list, 10)

    proveedores_page_number = request.GET.get('proveedores_page')
    productos_page_number = request.GET.get('productos_page')

    proveedores = proveedores_paginator.get_page(proveedores_page_number)
    productos = productos_paginator.get_page(productos_page_number)

    tipos_producto_distinct = Producto.objects.values_list('tipo_producto', flat=True).distinct()

    # --- CALCULO DE RANKING ---
    proveedores_ranking = []
    for p in Proveedor.objects.all():  # todos los proveedores, sin filtros
        desempenos = p.desempenos.all()
        if desempenos.exists():
            promedio_calidad = desempenos.aggregate(Avg('calidad'))['calidad__avg'] or 0
            promedio_entrega = desempenos.aggregate(Avg('entrega'))['entrega__avg'] or 0
            promedio_precio = desempenos.aggregate(Avg('precio'))['precio__avg'] or 0
            ranking = (promedio_calidad + promedio_entrega + promedio_precio) / 3
        else:
            ranking = 0
        p.ranking = round(ranking, 2)
        proveedores_ranking.append(p)

    # Orden descendente por ranking
    proveedores_ranking.sort(key=lambda x: x.ranking, reverse=True)

    # Paginación de ranking
    ranking_page_number = request.GET.get('ranking_page')
    ranking_paginator = Paginator(proveedores_ranking, 10)  
    proveedores_ranking_page = ranking_paginator.get_page(ranking_page_number)

    return render(request, 'proveedores/dashboard.html', {
        'proveedores': proveedores,
        'productos': productos,
        'tipos_producto': tipos_producto_distinct,
        'search_proveedor': search_proveedor,
        'orden': orden,
        'proveedores_ranking': proveedores_ranking_page,  
    })

def inicio(request):
    proveedores = Proveedor.objects.all()
    proveedores_con_ranking = []

    for p in proveedores:
        desempenos = p.desempenos.all()
        if desempenos.exists():
            promedio_calidad = desempenos.aggregate(Avg('calidad'))['calidad__avg'] or 0
            promedio_entrega = desempenos.aggregate(Avg('entrega'))['entrega__avg'] or 0
            promedio_precio = desempenos.aggregate(Avg('precio'))['precio__avg'] or 0
            ranking = (promedio_calidad + promedio_entrega + promedio_precio) / 3
        else:
            ranking = 0
        p.ranking = round(ranking, 2)
        proveedores_con_ranking.append(p)

    proveedores_con_ranking.sort(key=lambda x: x.ranking, reverse=True)

   
    page_number = request.GET.get('page', 1)
    paginator = Paginator(proveedores_con_ranking, 10)  # 10 por página
    proveedores_page = paginator.get_page(page_number)

    productos = Producto.objects.select_related('proveedor').all()
    tipos_producto = Producto.objects.values_list('tipo_producto', flat=True).distinct()

    tipo_id = request.GET.get('tipo_producto', '').strip()
    proveedor_id = request.GET.get('proveedor', '').strip()
    if tipo_id:
        productos = productos.filter(tipo_producto=tipo_id)
    if proveedor_id:
        try:
            pid = int(proveedor_id)
            productos = productos.filter(proveedor__id=pid)
        except ValueError:
            pass

    context = {
        'proveedores': proveedores_page,  # ahora es la página con ranking
        'productos': productos,
        'tipos_producto': tipos_producto,
    }
    return render(request, 'proveedores/inicio.html', context)

# Logout
def logout_view(request):
    logout(request)
    return redirect('login')
